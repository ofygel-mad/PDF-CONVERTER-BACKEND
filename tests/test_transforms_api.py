from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def test_parsers_endpoint_returns_supported_templates() -> None:
    client = TestClient(app)

    response = client.get("/api/v1/transforms/parsers")

    assert response.status_code == 200
    payload = response.json()
    assert any(item["key"] == "kaspi_gold_statement" for item in payload)
    assert any(item["key"] == "kaspi_business_statement" for item in payload)
    assert any(item["key"] == "ocr_scanned_statement" for item in payload)
    assert any(item["key"] == "generic_bank_statement" for item in payload)


def test_vision_status_endpoint_returns_runtime_state() -> None:
    client = TestClient(app)

    response = client.get("/api/v1/transforms/vision-status")

    assert response.status_code == 200
    payload = response.json()
    assert "available" in payload
    assert "backend" in payload
    assert "ocr_available" in payload
    assert "ocr_backend" in payload
    assert "use_cases" in payload


def test_preview_creates_history_record() -> None:
    client = TestClient(app)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Выписка Касpi Gold"
    rows = [
        ["ВЫПИСКА по Kaspi Gold за период с 08.03.26 по 20.03.26"],
        ["Клиент:", "Асан Бақдəулет Талғатұлы"],
        ["Номер карты:", "*1309"],
        ["Номер счета:", "KZ34722C000054174918"],
        ["Валюта счета:", "тенге"],
        ["Доступно на 08.03.26:", "100 426,14"],
        ["Доступно на 20.03.26:", "6 731,71"],
        [],
        ["Дата", "Приход, ₸", "Расход, ₸", "Детали / Операция"],
        ["20.03.26", 7000, None, "С Kaspi ДепозитаПополнение"],
    ]
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    preview = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("sample.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert preview.status_code == 200
    preview_payload = preview.json()
    assert "quality_summary" in preview_payload
    assert "row_diagnostics" in preview_payload
    assert preview_payload["quality_summary"]["overall_confidence"] <= 1.0

    history = client.get("/api/v1/transforms/history")

    assert history.status_code == 200
    payload = history.json()
    assert any(item["parser_key"] == "kaspi_gold_statement" for item in payload)


def test_template_creation_and_listing() -> None:
    client = TestClient(app)

    response = client.post(
        "/api/v1/transforms/templates",
        json={
            "parser_key": "kaspi_gold_statement",
            "name": "Финансист компакт",
            "description": "Только дата, расход и деталь",
            "base_variant_key": "classic_financier",
            "is_default": True,
            "columns": [
                {"key": "date", "label": "Дата", "kind": "text", "enabled": True},
                {"key": "income", "label": "Приход, ₸", "kind": "currency", "enabled": False},
                {"key": "expense", "label": "Расход, ₸", "kind": "currency", "enabled": True},
                {"key": "details_operation", "label": "Детали / Операция", "kind": "text", "enabled": True},
            ],
        },
    )

    assert response.status_code == 200
    template = response.json()
    assert template["is_default"] is True

    templates = client.get("/api/v1/transforms/templates?parser_key=kaspi_gold_statement")

    assert templates.status_code == 200
    assert any(item["template_id"] == template["template_id"] for item in templates.json())


def test_template_update_reorders_columns() -> None:
    client = TestClient(app)

    created = client.post(
        "/api/v1/transforms/templates",
        json={
            "parser_key": "kaspi_gold_statement",
            "name": "Проверка порядка",
            "description": "",
            "base_variant_key": "classic_financier",
            "is_default": False,
            "columns": [
                {"key": "date", "label": "Дата", "kind": "text", "enabled": True},
                {"key": "income", "label": "Приход, ₸", "kind": "currency", "enabled": True},
                {"key": "expense", "label": "Расход, ₸", "kind": "currency", "enabled": True},
            ],
        },
    ).json()

    response = client.patch(
        f"/api/v1/transforms/templates/{created['template_id']}",
        json={
            "name": "Проверка порядка v2",
            "columns": [
                {"key": "expense", "label": "Расход", "kind": "currency", "enabled": True},
                {"key": "date", "label": "Дата операции", "kind": "text", "enabled": True},
                {"key": "income", "label": "Приход", "kind": "currency", "enabled": False},
            ],
        },
    )

    assert response.status_code == 200
    updated = response.json()
    assert updated["name"] == "Проверка порядка v2"
    assert updated["columns"][0]["key"] == "expense"
    assert updated["columns"][1]["label"] == "Дата операции"
    assert updated["columns"][2]["enabled"] is False


def test_generic_bank_statement_autodetects_excel_table() -> None:
    client = TestClient(app)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Statement"
    rows = [
        ["Business Statement March"],
        [],
        ["Date", "Amount", "Description", "Operation"],
        ["2026-03-20", 12500, "Client payment", "Incoming transfer"],
        ["2026-03-21", -4800, "Office rent", "Outgoing transfer"],
    ]
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("generic.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "generic_bank_statement"
    assert any(item["key"] == "generic_bank_statement" and item["matched"] for item in payload["parser_matches"])
    assert payload["quality_summary"]["review_required_count"] >= 1
    assert any(item["flags"] for item in payload["row_diagnostics"])


def test_kaspi_business_statement_preview_builds_compact_variant() -> None:
    client = TestClient(app)
    workbook = Workbook()
    sheet = workbook.active
    rows = [
        ["Текущий счет:", None, "KZ68722S000009367266"],
        ["Валюта счета:", None, "KZT"],
        ["Период:", None, "03.04.2026 - 05.04.2026"],
        ["Дата последнего движения:", None, "03.04.2026 23:11"],
        ["ИИН/БИН:", None, "000305500863"],
        ["Наименование:", None, "ИП UMMA TENDER ACADEMY"],
        ["Входящий остаток:", None, 14770684.5],
        ["Исходящий остаток:", None, 24047202.82],
        [
            "№\nдокумента",
            "Дата операции",
            "Дебет",
            "Кредит",
            "Наименование бенефициара / отправителя денег",
            "ИИК бенефициара / отправителя денег",
            "БИК банка бенефициара (отправителя денег)",
            "КНП",
            "Назначение платежа",
        ],
        [1, 2, 3, 4, 5, 6, 7, 8, 9],
        [
            "10232853",
            "03.04.2026 21:40:05",
            500000,
            None,
            "Мийрибек Азизулы У.",
            "KZ31722C000032477139",
            "",
            "342",
            "Перевод собственных средств на карту Kaspi Gold *3205",
        ],
        [
            "10984841",
            "03.04.2026 23:11:36",
            None,
            12109000,
            'АО "KASPI BANK"\nИИН/БИН 971240001315',
            "KZ78722S000009367280",
            "",
            "190",
            "Продажи с Kaspi.kz за 03/04/2026",
        ],
    ]
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("kaspi-business.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "kaspi_business_statement"
    assert any(item["key"] == "kaspi_business_statement" and item["matched"] for item in payload["parser_matches"])

    compact_variant = next(item for item in payload["variants"] if item["key"] == "business_compact_classic")
    assert compact_variant["group"] == "kaspi_business_plus"
    assert compact_variant["rows"][0]["document_number"] == "10232853"
    assert compact_variant["rows"][0]["date"] == "03.04.2026"
    assert compact_variant["rows"][0]["expense"] == 500000
    assert compact_variant["rows"][0]["detail"] == "Мийрибек Азизулы У."
    assert compact_variant["rows"][0]["comment"] == "Перевод собственных средств на карту Kaspi Gold *3205"
    assert compact_variant["rows"][1]["income"] == 12109000
    assert compact_variant["rows"][1]["detail"] == 'АО "KASPI BANK"'


def test_scanned_image_statement_uses_ocr_parser(monkeypatch) -> None:
    from app.schemas.statement import ParsedStatement, StatementMetadata, StatementTotals, StatementTransaction

    client = TestClient(app)

    def fake_parse(filename: str, content: bytes) -> ParsedStatement:
        return ParsedStatement(
            metadata=StatementMetadata(
                source_filename=filename,
                title="Scanned statement",
                parser_key="ocr_scanned_statement",
                transaction_count=1,
                totals=StatementTotals(income_total=1500),
            ),
            transactions=[
                StatementTransaction(
                    date="05.04.26",
                    amount=1500,
                    income=1500,
                    expense=None,
                    operation="Incoming transfer",
                    detail="Client payment",
                    details_operation="Client payment Incoming transfer",
                    direction="inflow",
                )
            ],
        )

    monkeypatch.setattr("app.services.document_service._parse_ocr_statement", fake_parse)

    response = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("scan.jpeg", b"fake-image-bytes", "image/jpeg")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "ocr_scanned_statement"
    assert any(item["key"] == "ocr_scanned_statement" and item["matched"] for item in payload["parser_matches"])


def test_preview_falls_back_to_ocr_raw_review(monkeypatch) -> None:
    from app.schemas.statement import OCRReviewField, OCRReviewPayload, OCRReviewTable
    from app.services.document_service import DocumentParseError

    client = TestClient(app)

    def fake_preview(*_args, **_kwargs):
        raise DocumentParseError("No supported statement structure.")

    def fake_review(*_args, **_kwargs):
        return OCRReviewPayload(
            review_id="review-1",
            source_filename="scan.jpeg",
            lines=["raw one", "raw two"],
            tables=[
                OCRReviewTable(
                    table_index=0,
                    rows=[["Date", "Amount", "Description"], ["2026-04-05", "1500", "Payment"]],
                    suggested_header_row_index=0,
                )
            ],
            suggested_table_index=0,
            suggested_header_row_index=0,
            available_fields=[
                OCRReviewField(key="date", label="Date", required=True),
                OCRReviewField(key="detail", label="Detail", required=True),
                OCRReviewField(key="amount", label="Amount", required=True),
            ],
        )

    monkeypatch.setattr("app.api.routes.transforms.parse_statement_with_diagnostics", fake_preview)
    monkeypatch.setattr("app.api.routes.transforms.create_ocr_review_session", fake_review)

    response = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("scan.jpeg", b"fake-image-bytes", "image/jpeg")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "ocr_raw_review"
    assert payload["ocr_review"]["review_id"] == "review-1"
    assert payload["ocr_review"]["tables"][0]["rows"][0][0] == "Date"


def test_materialize_ocr_review_returns_normal_preview(monkeypatch) -> None:
    from app.schemas.statement import ParsedStatement, StatementMetadata, StatementTotals, StatementTransaction

    client = TestClient(app)
    saved = {"called": False}

    def fake_materialize(*_args, **_kwargs):
        return ParsedStatement(
            metadata=StatementMetadata(
                source_filename="scan.jpeg",
                title="Recovered statement",
                parser_key="ocr_scanned_statement",
                transaction_count=1,
                totals=StatementTotals(income_total=1200),
            ),
            transactions=[
                StatementTransaction(
                    date="05.04.26",
                    amount=1200,
                    income=1200,
                    expense=None,
                    operation="Incoming transfer",
                    detail="Recovered client",
                    details_operation="Recovered client Incoming transfer",
                    direction="inflow",
                    source="ocr",
                    source_confidence=0.7,
                )
            ],
        )

    def fake_save_template(*_args, **_kwargs):
        saved["called"] = True

    monkeypatch.setattr("app.api.routes.transforms.materialize_ocr_review", fake_materialize)
    monkeypatch.setattr("app.api.routes.transforms.save_mapping_template_from_review", fake_save_template)

    response = client.post(
        "/api/v1/transforms/ocr-reviews/review-1/materialize",
        json={
            "table_index": 0,
            "header_row_index": 0,
            "title": "Recovered statement",
            "save_mapping_template": True,
            "mapping_template_name": "Kaspi OCR Layout",
            "column_mapping": {"date": 0, "detail": 2, "amount": 1},
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "ocr_scanned_statement"
    assert payload["ocr_review"] is None
    assert payload["variants"]
    assert saved["called"] is True


def test_preview_auto_applies_saved_ocr_mapping_template(monkeypatch) -> None:
    from app.schemas.statement import (
        OCRMappingTemplate,
        OCRMappingTemplateMatch,
        OCRReviewField,
        OCRReviewPayload,
        OCRReviewTable,
        ParsedStatement,
        StatementMetadata,
        StatementTotals,
        StatementTransaction,
    )
    from app.services.document_service import DocumentParseError

    client = TestClient(app)

    def fake_preview(*_args, **_kwargs):
        raise DocumentParseError("No supported statement structure.")

    def fake_review(*_args, **_kwargs):
        return OCRReviewPayload(
            review_id="review-auto",
            source_filename="scan.jpeg",
            lines=["raw one"],
            tables=[
                OCRReviewTable(
                    table_index=0,
                    rows=[["Date", "Amount", "Description"], ["2026-04-05", "1500", "Payment"]],
                    suggested_header_row_index=0,
                )
            ],
            suggested_table_index=0,
            suggested_header_row_index=0,
            available_fields=[
                OCRReviewField(key="date", label="Date", required=True),
                OCRReviewField(key="detail", label="Detail", required=True),
                OCRReviewField(key="amount", label="Amount", required=True),
            ],
        )

    def fake_match(*_args, **_kwargs):
        return OCRMappingTemplateMatch(
            template_id="tpl-auto",
            name="Kaspi OCR Layout",
            score=0.91,
            table_index=0,
            header_row_index=0,
        )

    def fake_get_template(*_args, **_kwargs):
        return OCRMappingTemplate(
            template_id="tpl-auto",
            name="Kaspi OCR Layout",
            header_signature=["date", "amount", "description"],
            title_keywords=["kaspi"],
            column_mapping={"date": 0, "detail": 2, "amount": 1},
        )

    def fake_materialize(*_args, **_kwargs):
        return ParsedStatement(
            metadata=StatementMetadata(
                source_filename="scan.jpeg",
                title="Recovered by template",
                parser_key="ocr_scanned_statement",
                transaction_count=1,
                totals=StatementTotals(income_total=1500),
            ),
            transactions=[
                StatementTransaction(
                    date="05.04.26",
                    amount=1500,
                    income=1500,
                    expense=None,
                    operation="Incoming transfer",
                    detail="Payment",
                    details_operation="Payment Incoming transfer",
                    direction="inflow",
                    source="ocr",
                    source_confidence=0.7,
                )
            ],
        )

    monkeypatch.setattr("app.api.routes.transforms.parse_statement_with_diagnostics", fake_preview)
    monkeypatch.setattr("app.api.routes.transforms.create_ocr_review_session", fake_review)
    monkeypatch.setattr("app.api.routes.transforms.find_best_ocr_mapping_match", fake_match)
    monkeypatch.setattr("app.api.routes.transforms.get_ocr_mapping_template", fake_get_template)
    monkeypatch.setattr("app.api.routes.transforms.materialize_ocr_review", fake_materialize)

    response = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("scan.jpeg", b"fake-image-bytes", "image/jpeg")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["parser_key"] == "ocr_scanned_statement"
    assert payload["ocr_review"] is None
    assert payload["parser_matches"][0]["label"].startswith("OCR Auto Mapping")
    assert payload["applied_rule"]["name"] == "Kaspi OCR Layout"
    assert payload["applied_rule"]["version"] == 1
    assert payload["applied_rule"]["score"] == 0.91


def test_session_row_correction_updates_preview() -> None:
    client = TestClient(app)
    workbook = Workbook()
    sheet = workbook.active
    rows = [
        ["Business Statement March"],
        [],
        ["Date", "Amount", "Description", "Operation"],
        ["2026-03-20", 12500, "Client payment", "Incoming transfer"],
        ["2026-03-21", -4800, "Office rent", "Operation"],
    ]
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    preview = client.post(
        "/api/v1/transforms/preview",
        files={"file": ("generic.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    payload = preview.json()
    session_id = payload["session_id"]
    target_row = next(item for item in payload["row_diagnostics"] if item["row_number"] == 2)
    assert target_row["flags"]

    corrected = client.patch(
        f"/api/v1/transforms/sessions/{session_id}/rows/2",
        json={
            "operation": "Outgoing transfer",
            "detail": "Office rent LLC",
            "amount": 4800,
            "direction": "outflow",
        },
    )

    assert corrected.status_code == 200
    corrected_payload = corrected.json()
    corrected_row = next(item for item in corrected_payload["row_diagnostics"] if item["row_number"] == 2)
    assert corrected_row["source"] == "manual_correction"
    assert corrected_row["corrected"] is True
    assert corrected_payload["quality_summary"]["corrected_count"] >= 1
