from io import BytesIO

from openpyxl import load_workbook

from app.schemas.statement import ParsedStatement, StatementMetadata, StatementTotals, StatementTransaction
from app.services.export_service import export_statement


def _build_kaspi_gold_statement() -> ParsedStatement:
    return ParsedStatement(
        metadata=StatementMetadata(
            source_filename="sample.pdf",
            title="Выписка по Kaspi Gold за период с 14.04.26 по 15.04.26",
            parser_key="kaspi_gold_statement",
            account_holder="Муратов Калдыбай",
            card_number="*2550",
            account_number="KZ95722C000063945018",
            currency="тенге",
            opening_balance=1386112.17,
            closing_balance=1423768.17,
            totals=StatementTotals(
                topup_total=108300,
                transfer_total=45530,
                purchase_total=20644,
                cash_withdrawal_total=0,
            ),
        ),
        transactions=[
            StatementTransaction(
                date="15.04.26",
                amount=-25000,
                income=None,
                expense=25000,
                operation="Перевод",
                detail="Асхат К.",
                details_operation="Асхат К.Перевод",
                direction="outflow",
            ),
            StatementTransaction(
                date="15.04.26",
                amount=20000,
                income=20000,
                expense=None,
                operation="Пополнение",
                detail="Гульмира М.",
                details_operation="Гульмира М.Пополнение",
                direction="inflow",
            ),
        ],
    )


def test_export_statement_adds_operations_sheet_for_kaspi_gold() -> None:
    workbook_bytes = export_statement(_build_kaspi_gold_statement(), "classic_financier")

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert workbook.sheetnames == ["Выписка", "Операций", "Audit Trail"]

    operations_sheet = workbook["Операций"]
    headers = [operations_sheet.cell(9, column).value for column in range(1, 6)]
    assert headers == ["Дата", "Операция", "Детали", "Приход, ₸", "Расход, ₸"]
    assert operations_sheet.max_column == 5
    assert operations_sheet.freeze_panes == "B10"
    assert operations_sheet.cell(10, 2).value == "Перевод"
    assert operations_sheet.cell(11, 4).value == 20000


def test_export_operation_split_does_not_include_net_column() -> None:
    workbook_bytes = export_statement(_build_kaspi_gold_statement(), "operation_split")

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert workbook.sheetnames == ["Операций", "Audit Trail"]
    headers = [workbook["Операций"].cell(9, column).value for column in range(1, 6)]
    assert headers == ["Дата", "Операция", "Детали", "Приход, ₸", "Расход, ₸"]
