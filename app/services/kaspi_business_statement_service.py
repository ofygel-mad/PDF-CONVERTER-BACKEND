from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.statement import ParsedStatement, StatementMetadata, StatementTotals, StatementTransaction

ACCOUNT_LABEL = "\u0442\u0435\u043a\u0443\u0449\u0438\u0439 \u0441\u0447\u0435\u0442"
CURRENCY_LABEL = "\u0432\u0430\u043b\u044e\u0442\u0430 \u0441\u0447\u0435\u0442\u0430"
PERIOD_LABEL = "\u043f\u0435\u0440\u0438\u043e\u0434"
LAST_MOVEMENT_LABEL = (
    "\u0434\u0430\u0442\u0430 \u043f\u043e\u0441\u043b\u0435\u0434\u043d\u0435\u0433\u043e "
    "\u0434\u0432\u0438\u0436\u0435\u043d\u0438\u044f"
)
TAX_ID_LABEL = "\u0438\u0438\u043d/\u0431\u0438\u043d"
NAME_LABEL = "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
OPENING_LABEL = (
    "\u0432\u0445\u043e\u0434\u044f\u0449\u0438\u0439 \u043e\u0441\u0442\u0430\u0442\u043e\u043a"
)
CLOSING_LABEL = (
    "\u0438\u0441\u0445\u043e\u0434\u044f\u0449\u0438\u0439 \u043e\u0441\u0442\u0430\u0442\u043e\u043a"
)
DOCUMENT_HEADER = "\u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442"
DATE_HEADER = "\u0434\u0430\u0442\u0430 \u043e\u043f\u0435\u0440\u0430\u0446\u0438\u0438"
DEBIT_HEADER = "\u0434\u0435\u0431\u0435\u0442"
CREDIT_HEADER = "\u043a\u0440\u0435\u0434\u0438\u0442"
COUNTERPARTY_HEADER = (
    "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 "
    "\u0431\u0435\u043d\u0435\u0444\u0438\u0446\u0438\u0430\u0440\u0430"
)
COMMENT_HEADER = "\u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d\u0438\u0435 \u043f\u043b\u0430\u0442\u0435\u0436\u0430"
SELF_TRANSFER_PREFIX = (
    "\u043f\u0435\u0440\u0435\u0432\u043e\u0434 \u0441\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0445 "
    "\u0441\u0440\u0435\u0434\u0441\u0442\u0432 \u043d\u0430 \u043a\u0430\u0440\u0442\u0443 kaspi gold"
)
KASPI_SALES_PREFIX = "\u043f\u0440\u043e\u0434\u0430\u0436\u0438 \u0441 kaspi.kz"
KASPI_REFUND_PREFIX = (
    "\u0432\u043e\u0437\u0432\u0440\u0430\u0442 \u043f\u0440\u043e\u0434\u0430\u0436 \u0441 kaspi.kz"
)
PROCESSING_FEE_MARKER = (
    "\u043f\u0440\u043e\u0446\u0435\u0441\u0441\u0438\u043d\u0433\u0430"
)
MARKETPLACE_FEE_MARKER = (
    "\u043e\u0431\u0440\u0430\u0431\u043e\u0442\u043a\u0435 \u0434\u0430\u043d\u043d\u044b\u0445"
)
DEPOSIT_MARKER = "\u043d\u0430 \u0434\u0435\u043f\u043e\u0437\u0438\u0442"
OWN_ACCOUNT_MARKER = (
    "\u043d\u0430 \u0441\u0447\u0435\u0442 kaspipay"
)
TRANSFER_MARKER = "\u043f\u0435\u0440\u0435\u0432\u043e\u0434"
COUNTERPARTY_TYPE_PERSON = (
    "\u0424\u0438\u0437\u043b\u0438\u0446\u043e"
)
COUNTERPARTY_TYPE_BANK = "\u0411\u0430\u043d\u043a"
COUNTERPARTY_TYPE_SERVICE = "\u0421\u0435\u0440\u0432\u0438\u0441"
COUNTERPARTY_TYPE_BUSINESS = "\u0411\u0438\u0437\u043d\u0435\u0441"
FLOW_GROUP_SELF = "\u0421\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0435 \u043f\u0435\u0440\u0435\u0432\u043e\u0434\u044b"
FLOW_GROUP_REVENUE = "\u041f\u0440\u043e\u0434\u0430\u0436\u0438"
FLOW_GROUP_FEES = "\u041a\u043e\u043c\u0438\u0441\u0441\u0438\u0438"
FLOW_GROUP_DEPOSIT = "\u0414\u0435\u043f\u043e\u0437\u0438\u0442\u044b"
FLOW_GROUP_BANK = "\u0411\u0430\u043d\u043a\u043e\u0432\u0441\u043a\u0438\u0435 \u043e\u043f\u0435\u0440\u0430\u0446\u0438\u0438"
FLOW_GROUP_OTHER = "\u041f\u0440\u043e\u0447\u0435\u0435"

_TAX_ID_PATTERN = re.compile(
    r"(?:\s*\|\s*|\s+)?(?:\u0438\u0438\u043d|\u0431\u0438\u043d|\u0438\u0438\u043d/\u0431\u0438\u043d)"
    r"\s*[:№]?\s*\d{8,16}.*$",
    re.IGNORECASE,
)


def detect_kaspi_business_statement(content: bytes) -> float:
    rows = _load_rows(content, read_only=True, max_rows=20)
    header_index = _find_header_row(rows)
    if header_index is None:
        return 0.0

    score = 0.35
    for label in (ACCOUNT_LABEL, NAME_LABEL, PERIOD_LABEL):
        if _find_value_by_label(rows, label):
            score += 0.18
    if _find_value_by_label(rows, TAX_ID_LABEL):
        score += 0.07
    if _find_value_by_label(rows, LAST_MOVEMENT_LABEL):
        score += 0.08
    return min(0.99, round(score, 4))


def parse_kaspi_business_statement(filename: str, content: bytes) -> ParsedStatement:
    rows = _load_rows(content, read_only=False)
    header_index = _find_header_row(rows)
    if header_index is None:
        raise ValueError(
            "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043d\u0430\u0439\u0442\u0438 "
            "\u0442\u0430\u0431\u043b\u0438\u0446\u0443 \u043e\u043f\u0435\u0440\u0430\u0446\u0438\u0439 Kaspi Business."
        )

    column_map = _build_column_map(rows[header_index])
    transactions = _extract_transactions(rows[header_index + 1 :], column_map)
    if not transactions:
        raise ValueError(
            "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u0438\u0437\u0432\u043b\u0435\u0447\u044c "
            "\u0441\u0442\u0440\u043e\u043a\u0438 \u0438\u0437 \u0432\u044b\u043f\u0438\u0441\u043a\u0438 Kaspi Business."
        )

    metadata = _build_metadata(filename, rows, transactions)
    return ParsedStatement(metadata=metadata, transactions=transactions)


def _load_rows(
    content: bytes,
    *,
    read_only: bool,
    max_rows: int | None = None,
) -> list[list[object]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=read_only)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[list[object]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if max_rows is not None and index >= max_rows:
            break
    return rows


def _find_header_row(rows: list[list[object]]) -> int | None:
    for index, row in enumerate(rows):
        normalized = [_normalize_text(cell).lower() for cell in row if cell not in (None, "")]
        if not normalized:
            continue
        if (
            any(DOCUMENT_HEADER in cell for cell in normalized)
            and any(DATE_HEADER in cell for cell in normalized)
            and any(cell == DEBIT_HEADER for cell in normalized)
            and any(cell == CREDIT_HEADER for cell in normalized)
            and any(COMMENT_HEADER in cell for cell in normalized)
        ):
            return index
    return None


def _build_column_map(row: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(row):
        value = _normalize_text(cell).lower()
        if not value:
            continue
        if DOCUMENT_HEADER in value:
            mapping["document_number"] = index
        elif DATE_HEADER in value:
            mapping["operation_datetime"] = index
        elif value == DEBIT_HEADER:
            mapping["expense"] = index
        elif value == CREDIT_HEADER:
            mapping["income"] = index
        elif COUNTERPARTY_HEADER in value or (
            "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435" in value
            and "\u043e\u0442\u043f\u0440\u0430\u0432\u0438\u0442\u0435\u043b\u044f \u0434\u0435\u043d\u0435\u0433" in value
        ):
            mapping["counterparty"] = index
        elif COMMENT_HEADER in value:
            mapping["comment"] = index
    return mapping


def _extract_transactions(
    rows: list[list[object]],
    column_map: dict[str, int],
) -> list[StatementTransaction]:
    transactions: list[StatementTransaction] = []

    for row in rows:
        operation_date, operation_datetime = _coerce_operation_datetime(
            _cell(row, column_map.get("operation_datetime"))
        )
        if operation_date is None:
            continue

        income = _coerce_amount(_cell(row, column_map.get("income")))
        expense = _coerce_amount(_cell(row, column_map.get("expense")))
        if income is None and expense is None:
            continue

        raw_counterparty = _extract_counterparty(row, column_map)
        comment = _normalize_text(_cell(row, column_map.get("comment")))
        counterparty = _clean_counterparty(raw_counterparty, comment)
        operation = _classify_operation(comment, income=income, expense=expense)
        amount = income if income is not None else -(expense or 0.0)
        document_number = _normalize_text(_cell(row, column_map.get("document_number"))) or None

        transactions.append(
            StatementTransaction(
                date=operation_date,
                amount=round(amount, 2),
                income=round(income, 2) if income is not None else None,
                expense=round(expense, 2) if expense is not None else None,
                operation=operation,
                detail=counterparty or operation,
                details_operation=_compose_details_operation(counterparty, comment),
                direction="inflow" if income is not None else "outflow",
                document_number=document_number,
                operation_datetime=operation_datetime,
                comment=comment or None,
                raw_counterparty=raw_counterparty or None,
                source_cells={
                    "document_number": document_number or "",
                    "operation_datetime": operation_datetime or "",
                    "comment": comment,
                },
            )
        )

    return transactions


def _build_metadata(
    filename: str,
    rows: list[list[object]],
    transactions: list[StatementTransaction],
) -> StatementMetadata:
    account_number = _find_value_by_label(rows, ACCOUNT_LABEL)
    currency = _find_value_by_label(rows, CURRENCY_LABEL)
    period = _find_value_by_label(rows, PERIOD_LABEL)
    account_holder = _find_value_by_label(rows, NAME_LABEL)
    opening_balance = _coerce_amount(_find_value_by_label(rows, OPENING_LABEL))
    closing_balance = _coerce_amount(_find_value_by_label(rows, CLOSING_LABEL))
    period_start, period_end = _split_period(period)
    title = (
        f"\u0412\u044b\u043f\u0438\u0441\u043a\u0430 \u043f\u043e \u0441\u0447\u0435\u0442\u0443 {account_number}"
        if account_number
        else filename
    )

    return StatementMetadata(
        source_filename=filename,
        title=title,
        parser_key="kaspi_business_statement",
        account_holder=account_holder or None,
        account_number=account_number or None,
        currency=currency or None,
        period_start=period_start,
        period_end=period_end,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        transaction_count=len(transactions),
        totals=_calculate_totals(transactions),
    )


def _find_value_by_label(rows: list[list[object]], label: str) -> str:
    for row in rows:
        for index, cell in enumerate(row):
            normalized = _normalize_text(cell).lower().rstrip(":")
            if not normalized.startswith(label):
                continue
            for next_cell in row[index + 1 :]:
                value = _normalize_text(next_cell)
                if value:
                    return value
    return ""


def _coerce_operation_datetime(value: object | None) -> tuple[str | None, str | None]:
    if value in (None, ""):
        return None, None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y"), value.strftime("%d.%m.%Y %H:%M:%S")

    text = _normalize_text(value)
    if not text:
        return None, None

    for pattern in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.strftime("%d.%m.%Y"), parsed.strftime("%d.%m.%Y %H:%M:%S")
        except ValueError:
            continue
    return None, None


def _coerce_amount(value: object | None) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = _normalize_text(value)
    if not text:
        return None
    normalized = text.replace(" ", "").replace(",", ".")
    normalized = re.sub(r"[^\d\.\-+]", "", normalized)
    if not normalized:
        return None
    try:
        return round(float(Decimal(normalized)), 2)
    except InvalidOperation:
        return None


def _split_period(period: str) -> tuple[str | None, str | None]:
    if not period:
        return None, None
    match = re.match(r"^(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})$", period)
    if not match:
        return None, None
    return match.group(1), match.group(2)


def _clean_counterparty(raw_counterparty: str, comment: str) -> str:
    value = _normalize_text(raw_counterparty)
    if not value:
        return ""
    lowered_comment = comment.lower()
    if lowered_comment.startswith(SELF_TRANSFER_PREFIX):
        return value

    value = _TAX_ID_PATTERN.sub("", value)
    value = value.rstrip("|").strip()
    return value


def _classify_operation(comment: str, *, income: float | None, expense: float | None) -> str:
    comment_lower = comment.lower()
    if comment_lower.startswith(SELF_TRANSFER_PREFIX):
        return "\u0421\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0439 \u043f\u0435\u0440\u0435\u0432\u043e\u0434"
    if comment_lower.startswith(KASPI_REFUND_PREFIX):
        return "\u0412\u043e\u0437\u0432\u0440\u0430\u0442 Kaspi"
    if comment_lower.startswith(KASPI_SALES_PREFIX):
        return "\u041f\u0440\u043e\u0434\u0430\u0436\u0438 Kaspi"
    if PROCESSING_FEE_MARKER in comment_lower:
        return "\u041a\u043e\u043c\u0438\u0441\u0441\u0438\u044f \u043f\u0440\u043e\u0446\u0435\u0441\u0441\u0438\u043d\u0433\u0430"
    if MARKETPLACE_FEE_MARKER in comment_lower:
        return "\u041a\u043e\u043c\u0438\u0441\u0441\u0438\u044f \u043c\u0430\u0440\u043a\u0435\u0442\u043f\u043b\u0435\u0439\u0441\u0430"
    if DEPOSIT_MARKER in comment_lower:
        return "\u041f\u0435\u0440\u0435\u0432\u043e\u0434 \u043d\u0430 \u0434\u0435\u043f\u043e\u0437\u0438\u0442"
    if OWN_ACCOUNT_MARKER in comment_lower:
        return "\u041f\u0435\u0440\u0435\u0432\u043e\u0434 \u043d\u0430 \u0441\u0432\u043e\u0439 \u0441\u0447\u0435\u0442"
    if TRANSFER_MARKER in comment_lower:
        return "\u041f\u0435\u0440\u0435\u0432\u043e\u0434"
    if income is not None:
        return "\u041f\u043e\u0441\u0442\u0443\u043f\u043b\u0435\u043d\u0438\u0435"
    if expense is not None:
        return "\u0421\u043f\u0438\u0441\u0430\u043d\u0438\u0435"
    return "\u041e\u043f\u0435\u0440\u0430\u0446\u0438\u044f"


def _compose_details_operation(counterparty: str, comment: str) -> str:
    parts = [part for part in (counterparty, comment) if part]
    return " | ".join(parts)


def _extract_counterparty(row: list[object], column_map: dict[str, int]) -> str:
    explicit = _normalize_text(_cell(row, column_map.get("counterparty")))
    if explicit:
        return explicit

    ignored_indexes = {index for index in column_map.values()}
    for index, cell in enumerate(row):
        if index in ignored_indexes:
            continue
        text = _normalize_text(cell)
        if not text or text.isdigit():
            continue
        if re.search(r"[A-Za-z\u0400-\u04FF]", text):
            return text
    return ""


def derive_counterparty_type(detail: str) -> str:
    detail_upper = detail.upper()
    if "KASPI PAY" in detail_upper or "KASPI \u041c\u0410\u0413\u0410\u0417\u0418\u041d" in detail_upper:
        return COUNTERPARTY_TYPE_SERVICE
    if detail_upper.startswith("\u0410\u041e"):
        return COUNTERPARTY_TYPE_BANK
    if detail_upper.startswith("\u0422\u041e\u041e") or detail_upper.startswith("\u0418\u041f"):
        return COUNTERPARTY_TYPE_BUSINESS
    return COUNTERPARTY_TYPE_PERSON


def derive_flow_group(transaction: StatementTransaction) -> str:
    comment = (transaction.comment or "").lower()
    detail = (transaction.detail or "").lower()
    if comment.startswith(SELF_TRANSFER_PREFIX):
        return FLOW_GROUP_SELF
    if comment.startswith(KASPI_SALES_PREFIX) or comment.startswith(KASPI_REFUND_PREFIX):
        return FLOW_GROUP_REVENUE
    if PROCESSING_FEE_MARKER in comment or MARKETPLACE_FEE_MARKER in comment:
        return FLOW_GROUP_FEES
    if DEPOSIT_MARKER in comment:
        return FLOW_GROUP_DEPOSIT
    if "kaspi" in detail or "bank" in detail:
        return FLOW_GROUP_BANK
    return FLOW_GROUP_OTHER


def derive_flow_signal(transaction: StatementTransaction) -> str:
    if transaction.income is not None:
        return "\u0412\u0445\u043e\u0434\u044f\u0449\u0438\u0439 \u043f\u043e\u0442\u043e\u043a"
    return "\u0418\u0441\u0445\u043e\u0434\u044f\u0449\u0438\u0439 \u043f\u043e\u0442\u043e\u043a"


def derive_kaspi_bucket(transaction: StatementTransaction) -> str:
    comment = (transaction.comment or "").lower()
    if comment.startswith(KASPI_SALES_PREFIX):
        return "\u0412\u044b\u0440\u0443\u0447\u043a\u0430"
    if comment.startswith(KASPI_REFUND_PREFIX):
        return "\u0412\u043e\u0437\u0432\u0440\u0430\u0442"
    if PROCESSING_FEE_MARKER in comment:
        return "\u041f\u0440\u043e\u0446\u0435\u0441\u0441\u0438\u043d\u0433"
    if MARKETPLACE_FEE_MARKER in comment:
        return "Kaspi \u041c\u0430\u0433\u0430\u0437\u0438\u043d"
    if comment.startswith(SELF_TRANSFER_PREFIX):
        return "\u0412\u044b\u0432\u043e\u0434 \u043d\u0430 Kaspi Gold"
    if DEPOSIT_MARKER in comment:
        return "\u0414\u0435\u043f\u043e\u0437\u0438\u0442"
    return "\u041f\u0440\u043e\u0447\u0435\u0435"


def _calculate_totals(transactions: list[StatementTransaction]) -> StatementTotals:
    totals = defaultdict(float)
    for transaction in transactions:
        if transaction.income is not None:
            totals["income_total"] += transaction.income
        if transaction.expense is not None:
            totals["expense_total"] += transaction.expense
        if "\u043a\u043e\u043c\u0438\u0441\u0441\u0438\u044f" in transaction.operation.lower():
            totals["purchase_total"] += transaction.expense or 0.0
        if "\u043f\u0435\u0440\u0435\u0432\u043e\u0434" in transaction.operation.lower():
            totals["transfer_total"] += abs(transaction.amount)
        if transaction.income is not None:
            totals["topup_total"] += transaction.income
    return StatementTotals(**{key: round(value, 2) for key, value in totals.items()})


def _cell(row: list[object], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\n", " | ")
    text = unicodedata.normalize("NFC", text)
    return re.sub(r"[ \t]+", " ", text).strip()
