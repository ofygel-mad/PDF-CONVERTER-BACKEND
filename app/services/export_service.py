from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.statement import ParsedStatement
from app.services.template_service import get_template
from app.services.variant_service import apply_template_to_variant, build_variants

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_FILL_ODD = PatternFill(fill_type="solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill(fill_type="solid", fgColor="F6F9FC")
INFLOW_FILL = PatternFill(fill_type="solid", fgColor="E8F5E9")
OUTFLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF0F0")
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
SUBTLE_BORDER = Border(
    left=Side(style="thin", color="CAD5E2"),
    right=Side(style="thin", color="CAD5E2"),
    top=Side(style="thin", color="CAD5E2"),
    bottom=Side(style="thin", color="CAD5E2"),
)
TEXT_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")


def export_statement(
    statement: ParsedStatement,
    variant_key: str,
    excluded_rows: list[int] | None = None,
    custom_columns: list[dict] | None = None,
    custom_rows: list[dict] | None = None,
) -> bytes:
    from app.schemas.statement import PreviewColumn  # local import to avoid circular
    variants = {variant.key: variant for variant in build_variants(statement)}
    if variant_key.startswith("template::"):
        template_id = variant_key.split("template::", maxsplit=1)[1]
        template = get_template(template_id)
        if template is None:
            raise ValueError("Шаблон для экспорта не найден.")
        base_variant = variants.get(template.base_variant_key)
        if base_variant is None:
            raise ValueError("Базовый вариант шаблона не найден.")
        variants[variant_key] = apply_template_to_variant(base_variant, template)
    if variant_key not in variants:
        raise ValueError("Неизвестный вариант таблицы.")

    variant = variants[variant_key]

    # Apply custom layout (cell edits, added rows, column renames from inline editor)
    if custom_columns is not None:
        variant = variant.model_copy(update={
            "columns": [PreviewColumn(key=c["key"], label=c["label"], kind=c.get("kind", "text")) for c in custom_columns]
        })
    if custom_rows is not None:
        variant = variant.model_copy(update={"rows": custom_rows})
    elif excluded_rows:
        excluded_set = set(excluded_rows)
        variant = variant.model_copy(update={
            "rows": [r for i, r in enumerate(variant.rows, start=1) if i not in excluded_set]
        })
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preview Export"
    audit_sheet = workbook.create_sheet("Audit Trail")

    include_metadata = statement.metadata.parser_key != "halyk_fiz_statement"
    if include_metadata:
        _write_metadata(sheet, statement, len(variant.columns))
        header_row = 9
    else:
        header_row = 1

    for column_index, column in enumerate(variant.columns, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=column.label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = SUBTLE_BORDER
        cell.alignment = HEADER_ALIGNMENT

    for row_index, row in enumerate(variant.rows, start=header_row + 1):
        base_fill = _base_row_fill(row_index - header_row)
        for column_index, column in enumerate(variant.columns, start=1):
            value = row.get(column.key)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.fill = _cell_fill(column.key, column.kind, value, row.get("direction"), base_fill)
            cell.border = SUBTLE_BORDER
            cell.alignment = _alignment_for_column(column.key, column.kind)

    _apply_column_widths(sheet, variant.columns, variant.rows)
    _apply_row_heights(sheet, variant.columns, variant.rows, header_row)
    sheet.freeze_panes = _freeze_panes_for_variant(variant.key, header_row)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(variant.columns))}{max(header_row, len(variant.rows) + header_row)}"
    )
    sheet.sheet_view.showGridLines = False

    _write_audit_sheet(audit_sheet, statement, variant)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_metadata(sheet, statement: ParsedStatement, column_count: int) -> None:
    metadata = statement.metadata
    sheet.merge_cells(f"A1:{get_column_letter(max(column_count, 5))}1")
    sheet["A1"] = metadata.title
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")

    top_rows = [
        ("Клиент:", metadata.account_holder, "Пополнения", metadata.totals.topup_total),
        ("Номер карты:", metadata.card_number, "Переводы", metadata.totals.transfer_total),
        ("Номер счета:", metadata.account_number, "Покупки", metadata.totals.purchase_total),
        ("Валюта счета:", metadata.currency, "Снятия", metadata.totals.cash_withdrawal_total),
        ("Доступно на старте:", metadata.opening_balance, None, None),
        ("Доступно на финише:", metadata.closing_balance, None, None),
    ]

    for row_number, (left_label, left_value, right_label, right_value) in enumerate(top_rows, start=2):
        sheet.cell(row=row_number, column=1, value=left_label).font = LABEL_FONT
        sheet.cell(row=row_number, column=2, value=left_value)
        sheet.cell(row=row_number, column=1).alignment = TEXT_WRAP_ALIGNMENT
        sheet.cell(row=row_number, column=2).alignment = TEXT_WRAP_ALIGNMENT
        if right_label is not None:
            sheet.cell(row=row_number, column=4, value=right_label)
            sheet.cell(row=row_number, column=4).alignment = TEXT_WRAP_ALIGNMENT
        if right_value is not None:
            sheet.cell(row=row_number, column=5, value=right_value)
            sheet.cell(row=row_number, column=5).alignment = NUMBER_ALIGNMENT


def export_statement_csv(statement: ParsedStatement, variant_key: str) -> bytes:
    """Return a UTF-8 CSV of the chosen variant (with BOM for Excel compatibility)."""
    variants = {variant.key: variant for variant in build_variants(statement)}
    if variant_key.startswith("template::"):
        template_id = variant_key.split("template::", maxsplit=1)[1]
        template = get_template(template_id)
        if template is None:
            raise ValueError("Шаблон для экспорта не найден.")
        base_variant = variants.get(template.base_variant_key)
        if base_variant is None:
            raise ValueError("Базовый вариант шаблона не найден.")
        variants[variant_key] = apply_template_to_variant(base_variant, template)
    if variant_key not in variants:
        raise ValueError("Неизвестный вариант таблицы.")

    variant = variants[variant_key]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([col.label for col in variant.columns])
    for row in variant.rows:
        writer.writerow([row.get(col.key, "") for col in variant.columns])

    # BOM for correct Cyrillic display in Excel on Windows
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _write_audit_sheet(sheet, statement: ParsedStatement, variant) -> None:
    headers = [
        "Export row",
        "Export column",
        "Export header",
        "Value",
        "Source row",
        "Provenance",
        "Confidence",
        "Correction",
        "Template transform",
    ]
    for column_index, label in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    transaction_lookup = {index: item for index, item in enumerate(statement.transactions, start=1)}
    row_pointer = 2
    for export_row_number, row in enumerate(variant.rows, start=1):
        source_row_number = row.get("_source_row_number")
        if isinstance(source_row_number, int):
            source_row = transaction_lookup.get(source_row_number)
        else:
            source_row = transaction_lookup.get(export_row_number)
        for export_column_number, column in enumerate(variant.columns, start=1):
            value = row.get(column.key)
            provenance = row.get("_provenance") or (source_row.source if source_row else "derived_variant")
            confidence = source_row.source_confidence if source_row else None
            correction = "yes" if source_row and source_row.corrected else "no"
            template_transform = "yes" if variant.template_id else "no"
            sheet.cell(row=row_pointer, column=1, value=export_row_number)
            sheet.cell(row=row_pointer, column=2, value=export_column_number)
            sheet.cell(row=row_pointer, column=3, value=column.label)
            sheet.cell(row=row_pointer, column=4, value=value)
            sheet.cell(row=row_pointer, column=5, value=source_row_number if source_row else None)
            sheet.cell(row=row_pointer, column=6, value=provenance)
            sheet.cell(row=row_pointer, column=7, value=confidence)
            sheet.cell(row=row_pointer, column=8, value=correction)
            sheet.cell(row=row_pointer, column=9, value=template_transform)
            row_pointer += 1

    for column_index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20


def _apply_column_widths(sheet, columns, rows) -> None:
    sample_rows = rows[:120]
    for column_index, column in enumerate(columns, start=1):
        preferred = _preferred_width(column.key)
        if preferred is not None:
            sheet.column_dimensions[get_column_letter(column_index)].width = preferred
            continue

        max_length = max(
            len(str(column.label)),
            *(len(str(row.get(column.key, ""))) for row in sample_rows),
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 38)


def _preferred_width(column_key: str) -> float | None:
    return {
        "date": 14,
        "document_number": 16,
        "income": 16,
        "expense": 16,
        "amount": 16,
        "flow_type": 18,
        "flow_group": 22,
        "bucket": 20,
        "detail": 28,
        "comment": 32,
        "counterparty_type": 18,
        "self_transfer": 14,
        "currency_op": 10,
        "processing_date": 14,
    }.get(column_key)


def _apply_row_heights(sheet, columns, rows, header_row: int) -> None:
    wrap_keys = {column.key for column in columns if column.key in {"detail", "comment", "details_operation"}}
    if not wrap_keys:
        return

    for row_index, row in enumerate(rows, start=header_row + 1):
        longest = max((len(str(row.get(key, "") or "")) for key in wrap_keys), default=0)
        if longest > 150:
            sheet.row_dimensions[row_index].height = 66
        elif longest > 90:
            sheet.row_dimensions[row_index].height = 48
        elif longest > 45:
            sheet.row_dimensions[row_index].height = 34


def _alignment_for_column(column_key: str, kind: str) -> Alignment:
    if kind == "currency":
        return NUMBER_ALIGNMENT
    if column_key in {"date", "document_number", "self_transfer"}:
        return CENTER_ALIGNMENT
    return TEXT_WRAP_ALIGNMENT


def _base_row_fill(display_row_number: int) -> PatternFill:
    return ROW_FILL_EVEN if display_row_number % 2 == 0 else ROW_FILL_ODD


def _cell_fill(
    column_key: str,
    kind: str,
    value: object,
    direction: str | None,
    base_fill: PatternFill,
) -> PatternFill:
    if value in (None, "", "-"):
        return base_fill

    if kind == "currency":
        if column_key == "income":
            return INFLOW_FILL
        if column_key == "expense":
            return OUTFLOW_FILL
        if column_key == "amount":
            if direction == "inflow":
                return INFLOW_FILL
            if direction == "outflow":
                return OUTFLOW_FILL

    return base_fill


def _freeze_panes_for_variant(variant_key: str, header_row: int) -> str | None:
    if variant_key.startswith("business_") or variant_key.startswith("halyk_"):
        return None
    return f"B{header_row + 1}"
