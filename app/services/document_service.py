from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Callable

import fitz
from openpyxl import load_workbook

from app.schemas.statement import (
    ParsedStatement,
    ParserDescriptor,
    ParserMatch,
    StatementMetadata,
    StatementTotals,
    StatementTransaction,
)
from app.services.kaspi_business_statement import (
    detect_kaspi_business_statement,
    parse_kaspi_business_statement,
)
from app.services.ocr_service import OCRProcessingError, parse_ocr_statement

DATE_PATTERN = re.compile(r"^\d{2}\.\d{2}\.\d{2}$")
AMOUNT_PATTERN = re.compile(r"^[+-]?\s*[\d\s]+,\d{2}\s*₸?$")
KNOWN_OPERATIONS = ("Пополнение", "Перевод", "Покупка", "Снятие", "Разное")


class DocumentParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParserDefinition:
    key: str
    label: str
    description: str
    accepted_extensions: tuple[str, ...]
    detect: Callable[[str, bytes], float]
    parse: Callable[[str, bytes], ParsedStatement]


def parse_statement(filename: str, content: bytes) -> ParsedStatement:
    statement, _ = parse_statement_with_diagnostics(filename, content)
    return statement


def parse_statement_with_diagnostics(
    filename: str,
    content: bytes,
) -> tuple[ParsedStatement, list[ParserMatch]]:
    extension = Path(filename).suffix.lower()
    matched_extension = [parser for parser in _registered_parsers() if extension in parser.accepted_extensions]
    if not matched_extension:
        raise DocumentParseError("Поддерживаются только PDF и Excel-файлы.")

    matches = [
        ParserMatch(
            key=parser.key,
            label=parser.label,
            score=round(parser.detect(filename, content), 4),
        )
        for parser in matched_extension
    ]
    matches.sort(key=lambda match: match.score, reverse=True)

    selected_match = next((match for match in matches if match.score > 0), None)
    if selected_match is not None:
        parser = next(item for item in matched_extension if item.key == selected_match.key)
        selected_match.matched = True
        return parser.parse(filename, content), matches

    supported_labels = ", ".join(parser.label for parser in matched_extension)
    raise DocumentParseError(f"Документ не подошёл ни под один из доступных шаблонов: {supported_labels}.")


def list_supported_parsers() -> list[ParserDescriptor]:
    return [
        ParserDescriptor(
            key=parser.key,
            label=parser.label,
            description=parser.description,
            accepted_extensions=list(parser.accepted_extensions),
        )
        for parser in _registered_parsers()
    ]


def _parse_pdf_statement(filename: str, content: bytes) -> ParsedStatement:
    document = fitz.open(stream=content, filetype="pdf")
    pages = [_normalize(page.get_text("text")) for page in document]
    full_text = "\n".join(pages)
    if "Kaspi Gold" not in full_text or "ВЫПИСКА" not in full_text:
        raise DocumentParseError("Пока реализован парсер для выписок Kaspi Gold.")

    lines = [line for line in (_normalize(line) for line in full_text.splitlines()) if line]
    transactions = _extract_transactions(lines)
    metadata = _extract_pdf_metadata(filename, lines, transactions)
    return ParsedStatement(metadata=metadata, transactions=transactions)


def _parse_workbook_statement(filename: str, content: bytes) -> ParsedStatement:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [
        [_normalize(value) if isinstance(value, str) else value for value in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    flat_text = "\n".join(str(value) for row in rows for value in row if value)
    if "Kaspi Gold" not in flat_text or "ВЫПИСКА" not in flat_text:
        raise DocumentParseError("Пока реализован парсер для выписок Kaspi Gold.")

    transactions = _extract_transactions_from_normalized_sheet(rows)
    metadata = _extract_workbook_metadata(filename, rows, transactions)
    return ParsedStatement(metadata=metadata, transactions=transactions)


def _registered_parsers() -> list[ParserDefinition]:
    return [
        ParserDefinition(
            key="kaspi_gold_statement",
            label="Kaspi Gold Statement",
            description="Выписки Kaspi Gold в PDF или Excel с последующей нормализацией под финансовый шаблон.",
            accepted_extensions=(".pdf", ".xlsx", ".xlsm"),
            detect=_detect_kaspi_statement,
            parse=_parse_kaspi_statement,
        ),
        ParserDefinition(
            key="kaspi_business_statement",
            label="Kaspi Business Statement",
            description="Structured Kaspi Business or Kaspi Pay account statements in Excel format.",
            accepted_extensions=(".xlsx", ".xlsm"),
            detect=_detect_kaspi_business_statement,
            parse=_parse_kaspi_business_statement,
        ),
        ParserDefinition(
            key="ocr_scanned_statement",
            label="OCR Scanned Statement",
            description="OCR parser for scanned PDFs and image statements with table recovery before normalization.",
            accepted_extensions=(".pdf", ".png", ".jpg", ".jpeg"),
            detect=_detect_ocr_statement,
            parse=_parse_ocr_statement,
        ),
        ParserDefinition(
            key="generic_bank_statement",
            label="Generic Bank Statement",
            description="Общий табличный парсер для Excel-выписок с колонками даты, сумм и описания операции.",
            accepted_extensions=(".xlsx", ".xlsm"),
            detect=_detect_generic_bank_statement,
            parse=_parse_generic_bank_statement,
        ),
    ]


def _parse_kaspi_statement(filename: str, content: bytes) -> ParsedStatement:
    extension = Path(filename).suffix.lower()
    if extension == ".pdf":
        return _parse_pdf_statement(filename, content)
    return _parse_workbook_statement(filename, content)


def _detect_kaspi_statement(filename: str, content: bytes) -> float:
    extension = Path(filename).suffix.lower()
    if extension == ".pdf":
        document = fitz.open(stream=content, filetype="pdf")
        sample_text = "\n".join(_normalize(document[index].get_text("text")) for index in range(min(2, document.page_count)))
        return 1.0 if "Kaspi Gold" in sample_text and "ВЫПИСКА" in sample_text else 0.0
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        sample_values: list[str] = []
        for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
            sample_values.extend(str(value) for value in row if value is not None)
        joined = "\n".join(sample_values)
        return 1.0 if "Kaspi Gold" in joined and "ВЫПИСКА" in joined else 0.0
    return 0.0


def _detect_ocr_statement(filename: str, content: bytes) -> float:
    extension = Path(filename).suffix.lower()
    if extension in {".png", ".jpg", ".jpeg"}:
        return 0.78
    if extension == ".pdf":
        document = fitz.open(stream=content, filetype="pdf")
        sample_text = "\n".join(_normalize(document[index].get_text("text")) for index in range(min(2, document.page_count)))
        return 0.45 if len(sample_text.strip()) < 80 else 0.0
    return 0.0


def _detect_kaspi_business_statement(filename: str, content: bytes) -> float:
    extension = Path(filename).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        return 0.0
    return detect_kaspi_business_statement(content)


def _parse_kaspi_business_statement(filename: str, content: bytes) -> ParsedStatement:
    try:
        return parse_kaspi_business_statement(filename, content)
    except ValueError as exc:
        raise DocumentParseError(str(exc)) from exc


def _parse_ocr_statement(filename: str, content: bytes) -> ParsedStatement:
    try:
        return parse_ocr_statement(filename, content)
    except OCRProcessingError as exc:
        raise DocumentParseError(str(exc)) from exc


def _detect_generic_bank_statement(filename: str, content: bytes) -> float:
    extension = Path(filename).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        return 0.0

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [
        [_normalize(value) if isinstance(value, str) else value for value in row]
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True)
    ]
    header_index = _find_generic_header_row(rows)
    if header_index is None:
        return 0.0

    column_map = _build_generic_column_map(rows[header_index])
    score = 0.0
    if "date" in column_map:
        score += 0.45
    if "detail" in column_map:
        score += 0.25
    if "amount" in column_map or ("income" in column_map and "expense" in column_map):
        score += 0.30
    return score


def _parse_generic_bank_statement(filename: str, content: bytes) -> ParsedStatement:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [
        [_normalize(value) if isinstance(value, str) else value for value in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    header_index = _find_generic_header_row(rows)
    if header_index is None:
        raise DocumentParseError("Не удалось найти заголовки табличной банковской выписки.")

    column_map = _build_generic_column_map(rows[header_index])
    transactions = _extract_generic_transactions(rows[header_index + 1 :], column_map)
    if not transactions:
        raise DocumentParseError("Не удалось извлечь строки операций из табличной выписки.")

    title = _normalize(str(rows[0][0])) if rows and rows[0] and rows[0][0] is not None else filename
    metadata = StatementMetadata(
        source_filename=filename,
        title=title,
        parser_key="generic_bank_statement",
        transaction_count=len(transactions),
        totals=_calculate_totals(transactions),
    )
    return ParsedStatement(metadata=metadata, transactions=transactions)


def _extract_pdf_metadata(
    filename: str,
    lines: list[str],
    transactions: list[StatementTransaction],
) -> StatementMetadata:
    title_line = next((line for line in lines if line.startswith("ВЫПИСКА")), "Выписка")
    if title_line == "ВЫПИСКА":
        title_index = lines.index(title_line)
        if title_index + 1 < len(lines) and "Kaspi Gold" in lines[title_index + 1]:
            title_line = f"{title_line} {lines[title_index + 1]}"
    period_match = re.search(r"с\s+(\d{2}\.\d{2}\.\d{2})\s+по\s+(\d{2}\.\d{2}\.\d{2})", title_line)
    card_index = lines.index("Номер карты:") if "Номер карты:" in lines else -1
    account_index = lines.index("Номер счета:") if "Номер счета:" in lines else -1

    account_holder = None
    card_number = None
    account_number = None
    if card_index > 0 and card_index + 2 < len(lines):
        card_number = lines[card_index + 1]
        account_holder = " ".join(part for part in [lines[card_index - 1], lines[card_index + 2]] if part)
    if account_index > -1 and account_index + 1 < len(lines):
        account_number = lines[account_index + 1]

    opening_label = next((line for line in lines if line.startswith("Доступно на ") and line.endswith(":")), None)
    opening_balance = None
    closing_balance = None
    if opening_label and opening_label in lines:
        opening_idx = lines.index(opening_label)
        if opening_idx + 1 < len(lines):
            opening_balance = _parse_amount(lines[opening_idx + 1])
    closing_label = next(
        (
            line
            for line in reversed(lines)
            if line.startswith("Доступно на ") and line.endswith(":") and line != opening_label
        ),
        None,
    )
    if closing_label and closing_label in lines:
        closing_idx = lines.index(closing_label)
        if closing_idx + 1 < len(lines):
            closing_balance = _parse_amount(lines[closing_idx + 1])

    currency = None
    if "Валюта счета:" in lines:
        currency_idx = lines.index("Валюта счета:")
        if currency_idx + 1 < len(lines):
            currency = lines[currency_idx + 1]

    totals = _calculate_totals(transactions)
    return StatementMetadata(
        source_filename=filename,
        title=title_line,
        parser_key="kaspi_gold_statement",
        account_holder=account_holder,
        card_number=card_number,
        account_number=account_number,
        currency=currency,
        period_start=period_match.group(1) if period_match else None,
        period_end=period_match.group(2) if period_match else None,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        transaction_count=len(transactions),
        totals=totals,
    )


def _extract_workbook_metadata(
    filename: str,
    rows: list[list[object]],
    transactions: list[StatementTransaction],
) -> StatementMetadata:
    title = str(rows[0][0]) if rows and rows[0] and rows[0][0] else "Выписка"
    period_match = re.search(r"с\s+(\d{2}\.\d{2}\.\d{2})\s+по\s+(\d{2}\.\d{2}\.\d{2})", title)
    account_holder = _string_or_none(rows, 1, 1)
    card_number = _string_or_none(rows, 2, 1)
    account_number = _string_or_none(rows, 3, 1)
    currency = _string_or_none(rows, 4, 1)
    opening_balance = _parse_amount(str(rows[5][1])) if _value_exists(rows, 5, 1) else None
    closing_balance = _parse_amount(str(rows[6][1])) if _value_exists(rows, 6, 1) else None

    return StatementMetadata(
        source_filename=filename,
        title=title,
        parser_key="kaspi_gold_statement",
        account_holder=account_holder,
        card_number=card_number,
        account_number=account_number,
        currency=currency,
        period_start=period_match.group(1) if period_match else None,
        period_end=period_match.group(2) if period_match else None,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        transaction_count=len(transactions),
        totals=_calculate_totals(transactions),
    )


def _extract_transactions(lines: list[str]) -> list[StatementTransaction]:
    try:
        header_index = lines.index("Дата")
    except ValueError as exc:
        raise DocumentParseError("Не удалось найти таблицу операций в PDF.") from exc

    body = lines[header_index + 3 :]
    transactions: list[StatementTransaction] = []
    index = 0

    while index < len(body):
        line = body[index]
        if not DATE_PATTERN.match(line):
            index += 1
            continue

        if index + 2 >= len(body):
            break

        amount_line = body[index + 1]
        operation_line = body[index + 2]
        if not AMOUNT_PATTERN.match(amount_line):
            index += 1
            continue

        operation, detail = _split_operation_detail(operation_line)
        note = None
        flags: list[str] = []

        next_index = index + 3
        while next_index < len(body) and not DATE_PATTERN.match(body[next_index]):
            candidate = body[next_index]
            if candidate.startswith("- ") or candidate.startswith("• ") or candidate.startswith("—"):
                note = candidate
                flags.append("requires_attention")
            next_index += 1

        transactions.append(
            _build_transaction(
                date=line,
                amount_line=amount_line,
                operation=operation,
                detail=detail,
                note=note,
                flags=flags,
            )
        )
        index = next_index

    if not transactions:
        raise DocumentParseError("Не удалось распарсить операции из PDF.")
    return transactions


def _extract_transactions_from_normalized_sheet(rows: list[list[object]]) -> list[StatementTransaction]:
    header_row = None
    for index, row in enumerate(rows):
        normalized = [str(cell).strip() if cell is not None else "" for cell in row]
        if "Дата" in normalized and "Детали / Операция" in normalized:
            header_row = index
            break

    if header_row is None:
        raise DocumentParseError("Не удалось найти строку заголовков в Excel.")

    transactions: list[StatementTransaction] = []
    for row in rows[header_row + 1 :]:
        if not row or row[0] in (None, ""):
            continue
        date = _normalize(str(row[0]))
        if not DATE_PATTERN.match(date):
            continue
        income = _as_float(row[1]) if len(row) > 1 else None
        expense = _as_float(row[2]) if len(row) > 2 else None
        combined = _normalize(str(row[3])) if len(row) > 3 and row[3] is not None else ""
        operation, detail = _split_combined_value(combined)
        amount = income if income is not None else -(expense or 0.0)
        transactions.append(
            StatementTransaction(
                date=date,
                amount=round(amount, 2),
                income=round(income, 2) if income is not None else None,
                expense=round(expense, 2) if expense is not None else None,
                operation=operation,
                detail=detail,
                details_operation=combined,
                direction="inflow" if income is not None else "outflow",
            )
        )

    if not transactions:
        raise DocumentParseError("Не удалось извлечь операции из Excel.")
    return transactions


def _find_generic_header_row(rows: list[list[object]]) -> int | None:
    for index, row in enumerate(rows):
        normalized = [_normalize(str(cell)).lower() if cell is not None else "" for cell in row]
        if "date" in normalized or "дата" in normalized:
            if any(token in normalized for token in ["amount", "сумма", "расход", "приход", "description", "описание", "детали", "details"]):
                return index
    return None


def _build_generic_column_map(row: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(row):
        value = _normalize(str(cell)).lower() if cell is not None else ""
        if value in {"date", "дата"}:
            mapping["date"] = index
        elif value in {"income", "credit", "inflow", "приход", "поступление", "зачисление"}:
            mapping["income"] = index
        elif value in {"expense", "debit", "outflow", "расход", "списание"}:
            mapping["expense"] = index
        elif value in {"amount", "сумма"}:
            mapping["amount"] = index
        elif value in {"description", "details", "detail", "описание", "детали", "назначение"}:
            mapping["detail"] = index
        elif value in {"operation", "type", "операция", "тип"}:
            mapping["operation"] = index
    return mapping


def _extract_generic_transactions(
    rows: list[list[object]],
    column_map: dict[str, int],
) -> list[StatementTransaction]:
    transactions: list[StatementTransaction] = []

    for row in rows:
        if "date" not in column_map:
            break
        date_value = _cell_value(row, column_map["date"])
        if not date_value:
            continue
        date = _coerce_date(date_value)
        if date is None:
            continue

        income = _coerce_numeric_cell(_cell_value(row, column_map.get("income")))
        expense = _coerce_numeric_cell(_cell_value(row, column_map.get("expense")))
        amount = _coerce_numeric_cell(_cell_value(row, column_map.get("amount")))

        if amount is not None and income is None and expense is None:
            if amount >= 0:
                income = amount
            else:
                expense = abs(amount)

        if income is None and expense is None:
            continue

        operation = _normalize(str(_cell_value(row, column_map.get("operation")) or "Операция"))
        detail = _normalize(str(_cell_value(row, column_map.get("detail")) or operation))
        signed_amount = income if income is not None else -(expense or 0.0)

        transactions.append(
            StatementTransaction(
                date=date,
                amount=round(signed_amount, 2),
                income=round(income, 2) if income is not None else None,
                expense=round(expense, 2) if expense is not None else None,
                operation=operation,
                detail=detail,
                details_operation=f"{detail}{operation}",
                direction="inflow" if income is not None else "outflow",
            )
        )

    return transactions


def _build_transaction(
    date: str,
    amount_line: str,
    operation: str,
    detail: str,
    note: str | None,
    flags: list[str],
) -> StatementTransaction:
    amount = _parse_amount(amount_line)
    direction = "inflow" if amount >= 0 else "outflow"
    income = round(amount, 2) if amount >= 0 else None
    expense = round(abs(amount), 2) if amount < 0 else None
    return StatementTransaction(
        date=date,
        amount=round(amount, 2),
        income=income,
        expense=expense,
        operation=operation,
        detail=detail,
        details_operation=f"{detail}{operation}",
        direction=direction,
        note=note,
        flags=flags,
    )


def _split_operation_detail(value: str) -> tuple[str, str]:
    match = re.match(r"^(?P<operation>.+?)\s{2,}(?P<detail>.+)$", value)
    if match:
        return _normalize(match.group("operation")), _normalize(match.group("detail"))
    for operation in KNOWN_OPERATIONS:
        if value.startswith(operation):
            return operation, _normalize(value[len(operation) :]) or operation
    return "Операция", value


def _split_combined_value(value: str) -> tuple[str, str]:
    for operation in sorted(KNOWN_OPERATIONS, key=len, reverse=True):
        if value.endswith(operation):
            detail = _normalize(value[: -len(operation)])
            return operation, detail or operation
    return "Операция", value


def _calculate_totals(transactions: list[StatementTransaction]) -> StatementTotals:
    totals = defaultdict(float)
    for transaction in transactions:
        if transaction.income is not None:
            totals["income_total"] += transaction.income
        if transaction.expense is not None:
            totals["expense_total"] += transaction.expense
        if transaction.operation == "Покупка" and transaction.expense is not None:
            totals["purchase_total"] += transaction.expense
        if transaction.operation == "Перевод" and transaction.expense is not None:
            totals["transfer_total"] += transaction.expense
        if transaction.operation == "Пополнение" and transaction.income is not None:
            totals["topup_total"] += transaction.income
        if transaction.operation == "Снятие" and transaction.expense is not None:
            totals["cash_withdrawal_total"] += transaction.expense

    return StatementTotals(**{key: round(value, 2) for key, value in totals.items()})


def _parse_amount(value: str) -> float:
    normalized = _normalize(value).replace("₸", "").replace(" ", "")
    sign = -1 if normalized.startswith("-") else 1
    normalized = normalized.lstrip("+-")
    normalized = normalized.replace(",", ".")
    try:
        amount = Decimal(normalized)
    except InvalidOperation as exc:
        raise DocumentParseError(f"Не удалось распарсить сумму: {value}") from exc
    return round(float(amount) * sign, 2)


def _as_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return _parse_amount(str(value))


def _cell_value(row: list[object], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _coerce_numeric_cell(value: object | None) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = _normalize(str(value))
    if not text:
        return None
    try:
        return _parse_amount(text)
    except DocumentParseError:
        return None


def _coerce_date(value: object) -> str | None:
    text = _normalize(str(value))
    if DATE_PATTERN.match(text):
        return text
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if iso_match:
        return f"{iso_match.group(3)}.{iso_match.group(2)}.{iso_match.group(1)[2:]}"
    slash_match = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", text)
    if slash_match:
        return f"{slash_match.group(1)}.{slash_match.group(2)}.{slash_match.group(3)[2:]}"
    return None


def _normalize(value: str) -> str:
    collapsed = value.replace("\xa0", " ").replace("−", "-")
    collapsed = unicodedata.normalize("NFC", collapsed)
    return re.sub(r"[ \t]+", " ", collapsed).strip()


def _string_or_none(rows: list[list[object]], row_index: int, column_index: int) -> str | None:
    if not _value_exists(rows, row_index, column_index):
        return None
    return _normalize(str(rows[row_index][column_index]))


def _value_exists(rows: list[list[object]], row_index: int, column_index: int) -> bool:
    return row_index < len(rows) and column_index < len(rows[row_index]) and rows[row_index][column_index] not in (None, "")
